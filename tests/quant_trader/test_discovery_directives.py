"""Tests for the excluded_tickers / preferred_lookback_days additions to
propose_pairs (mandate-directive support, mirroring Fundamental Trader's
propose_category_deviations - see mandate_directives.py).

Uses the real, local 120-ticker ETF_historical_prices.xlsx fixture rather
than a hand-built synthetic panel: constructing synthetic price series that
satisfy find_correlated_pairs' correlation threshold *and*
estimate_half_life's AR(1) mean-reversion fit (on the price *ratio*, not a
difference) turned out to need more tuning than it was worth, when a real,
known-good dataset is already sitting in the repo root for local
development. Skips cleanly if that file isn't present (e.g. in CI, or a
fresh clone before the data fixture has been copied in) rather than
failing - this is a supplementary check, not the only regression coverage
for these code paths: the full pipeline (agent -> backtest -> Risk) was
independently verified live against this same dataset during development.
"""

from __future__ import annotations

from datetime import timezone
from pathlib import Path

import pytest

REPO_ROOT = Path(__file__).resolve().parents[2]
PRICES_PATH = REPO_ROOT / "ETF_historical_prices.xlsx"

pytestmark = pytest.mark.skipif(
    not PRICES_PATH.exists(),
    reason="ETF_historical_prices.xlsx not present locally (gitignored fixture).",
)


def _load_offline_panel():
    from openpyxl import load_workbook

    from tools import PriceBar

    wb = load_workbook(str(PRICES_PATH), read_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    idx = {h: i for i, h in enumerate(header)}
    panel: dict[str, list] = {}
    for r in rows:
        ticker, dt, close = r[idx["ticker"]], r[idx["date"]], r[idx["close"]]
        if ticker is None or dt is None or close is None:
            continue
        panel.setdefault(ticker, []).append(PriceBar(
            symbol=ticker, timestamp=dt.replace(tzinfo=timezone.utc),
            open=r[idx["open"]] or close, high=r[idx["high"]] or close,
            low=r[idx["low"]] or close, close=close,
        ))
    return {k: tuple(v) for k, v in panel.items()}


@pytest.fixture(scope="module")
def offline_panel():
    return _load_offline_panel()


def test_excluded_tickers_removes_pairs_involving_that_symbol(offline_panel):
    from agents.quant_trader.discovery import propose_pairs

    baseline = propose_pairs(offline_panel, top_n=5)
    assert baseline, "expected at least one candidate pair from the real dataset"

    target = baseline[0].ticker_a
    excluded = propose_pairs(offline_panel, excluded_tickers={target}, top_n=5)

    assert all(target not in (p.ticker_a, p.ticker_b) for p in excluded)


def test_preferred_lookback_days_overrides_the_half_life_derived_value(offline_panel):
    from agents.quant_trader.discovery import propose_pairs

    baseline = propose_pairs(offline_panel, top_n=1)
    assert baseline

    overridden = propose_pairs(offline_panel, top_n=1, preferred_lookback_days=77)
    assert overridden
    assert overridden[0].lookback_days == 77


def test_entry_and_exit_zscore_are_passed_through_to_every_proposal(offline_panel):
    from agents.quant_trader.discovery import propose_pairs

    proposals = propose_pairs(offline_panel, top_n=5, entry_zscore=2.1, exit_zscore=0.4)

    assert proposals
    assert all(p.entry_zscore == 2.1 for p in proposals)
    assert all(p.exit_zscore == 0.4 for p in proposals)
