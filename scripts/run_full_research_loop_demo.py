#!/usr/bin/env python3
"""Full research-loop integration demo: PM -> 3 traders -> Risk -> Reporting -> PM -> Memory.

Runs the REAL compiled production graph (``graph.production.compile_production_workflow``)
with:

  * REAL Fundamental Trader  - agents.fundamental_trader (Aditi)
  * REAL Quant Trader        - agents.quant_trader (Shaurya) - no LLM required
  * REAL/STUBBED Technical Trader - see note below
  * REAL Risk Agent          - agents.risk_agent.RiskAgentImpl (Yutong)
  * REAL Reporting Agent     - agents.reporting_agent.ReportingAgentImpl (Emma)
  * REAL PM decision         - a durable LangGraph interrupt, resumable via --resume
  * REAL Memory              - services.file_memory_store.FileBackedMemoryStore

Technical Trader: real when credentials are configured, stubbed otherwise
----------------------------------------------------------------------------
TechnicalTraderAgent.run() calls a real ``ModelClient.generate_structured(...)``.
A concrete OpenAI/Anthropic implementation now exists
(``agents.technical_trader.adapters``), gated behind environment variables -
see ``agents/technical_trader/docs/integration.md``. This script constructs
the real runtime whenever ``TECHNICAL_TRADER_MODEL_PROVIDER`` (and the
matching model + API key) are set; otherwise it falls back to a clearly
labeled stub node so the graph topology still runs end-to-end.

The composition path supports both OpenAI and Anthropic. A missing provider
selection deliberately uses the labeled stub; once a provider is selected,
incomplete or invalid configuration fails clearly instead of silently
downgrading the run.

Everything downstream of the trader join (Risk, Reporting, PM decision,
Memory write) runs for real regardless of which trader nodes are real vs
stubbed - the graph doesn't know or care.

Why PM decisions use a durable interrupt, and Memory persists to disk
--------------------------------------------------------------------------
Portfolio Manager is deliberately the human's seat (see
``management/portfolio_manager.py``). The graph's own
``human_pm_decision_node`` (not overridden here) pauses on a real LangGraph
interrupt and waits for ``--resume --decision-json ...``. Because the
dashboard launches each round as a fresh subprocess, both the paused
interrupt and recorded Memory need to survive a process exiting - hence
``AsyncSqliteSaver`` (checkpoints) and ``FileBackedMemoryStore`` (Memory),
both backed by files under ``dashboard/data/``, instead of the in-process
``MemorySaver``/``InMemoryMemoryStore`` this script used previously.

Run (offline, no network for data/backtesting - uses the real 120-ticker
ETF fixtures; Technical Trader's own model call, if configured, does use
the network):

    python scripts/run_full_research_loop_demo.py
    python scripts/run_full_research_loop_demo.py --resume --run-id <id> --decision-json <path>
"""

from __future__ import annotations

import asyncio
import argparse
import json
import os
import sys
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any, Mapping

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from protocols import (  # noqa: E402
    BacktestRequest,
    DataArtifact,
    DataCategory,
    DataProvenance,
    DataRequest,
    DataResponse,
    PMDecision,
    PMDecisionType,
    PMMandate,
    RunStatus,
    SpecialistId,
    TaskLineage,
    TraderStrategyPackage,
)
from tools import DeterministicBacktestEngine, PriceBar, ResolvedBacktestData  # noqa: E402

from agents.fundamental_trader import (  # noqa: E402
    FundamentalTraderRuntime,
    category_deviation_executor,
)
from agents.fundamental_trader.agent import FundamentalTraderAgent  # noqa: E402
from agents.fundamental_trader.examples.static_data_service import _load_etf_metadata  # noqa: E402
from agents.fundamental_trader.examples.validation_split import (  # noqa: E402
    PercentileValidationSplitPolicy,
)
from agents.quant_trader import (  # noqa: E402
    QuantTraderAgent,
    QuantTraderRuntime,
    cross_asset_spread_executor,
)
from agents.risk_agent import RiskAgentImpl, make_risk_review_node  # noqa: E402
from agents.reporting_agent import ReportingAgentImpl  # noqa: E402
from protocols.reporting import ReportingRequest  # noqa: E402
from dashboard.workflow_adapter import write_dashboard_snapshot  # noqa: E402
from integration import WorkflowRunner  # noqa: E402
from services.file_memory_store import FileBackedMemoryStore  # noqa: E402
from protocols.research_contracts import MemoryRecord  # noqa: E402
from scripts.horizon_matched_validation import (  # noqa: E402
    HorizonMatchedValidationSplitPolicy,
)
from scripts.full_test_identity import derive_demo_identifiers  # noqa: E402

# Real Technical Trader is optional at import time: only constructed in main()
# if TECHNICAL_TRADER_MODEL_PROVIDER etc. are actually set (see
# _build_technical_trader_node below). Importing the package itself never
# reads credentials - see agents/technical_trader/docs/integration.md.
from agents.technical_trader import (  # noqa: E402
    ExecutionPolicy as TechnicalExecutionPolicy,
    JsonFileTechnicalDiagnosticsSink,
    TECHNICAL_STRATEGY_EXECUTORS,
    TechnicalModelConfigurationError,
    create_technical_model_client_from_env,
    create_technical_trader_runtime,
)

# A ticker present in the offline 120-ETF fixture, used as Technical Trader's
# PM-approved shared benchmark (see integration.md - required for its
# out-of-sample "must beat the benchmark" gate).
TECHNICAL_TRADER_BENCHMARK_SYMBOL = "IVV"

langgraph_import_error: ImportError | None = None
try:
    from graph.production import ProductionNodeSet, compile_production_workflow  # noqa: E402
    from langgraph.checkpoint.sqlite.aio import AsyncSqliteSaver  # noqa: E402
except ImportError as import_error:
    langgraph_import_error = import_error


REPO_ROOT = Path(__file__).resolve().parents[1]

# Loading a repo-local .env is a convenience for this executable composition
# root only. Shell/deployment variables retain precedence, and importing the
# Technical Trader package itself never reads credentials.
try:
    from dotenv import load_dotenv
except ImportError:  # Environment-only configuration remains supported.
    load_dotenv = None
if load_dotenv is not None:
    load_dotenv(REPO_ROOT / ".env", override=False)


def _offline_data_path(environment_name: str, default_name: str) -> Path:
    """Resolve an optional data path relative to the repository root."""

    configured = os.environ.get(environment_name, "").strip()
    path = Path(configured).expanduser() if configured else Path(default_name)
    if not path.is_absolute():
        path = REPO_ROOT / path
    return path.resolve()


OFFLINE_PRICES_PATH = _offline_data_path(
    "ETF_HISTORICAL_PRICES_PATH",
    "ETF_historical_prices.xlsx",
)
OFFLINE_METADATA_PATH = _offline_data_path("ETF_INFO_PATH", "ETF_info.xlsx")


def _require_offline_data() -> None:
    missing = [
        path
        for path in (OFFLINE_PRICES_PATH, OFFLINE_METADATA_PATH)
        if not path.is_file()
    ]
    if not missing:
        return
    formatted = "\n".join(f"  - {path}" for path in missing)
    raise FileNotFoundError(
        "The full-loop demo requires the team-supplied offline ETF data. "
        "The following file(s) were not found:\n"
        f"{formatted}\n"
        "Place the workbooks in the repository root or set "
        "ETF_HISTORICAL_PRICES_PATH and ETF_INFO_PATH in .env."
    )


RUN_ID, WORKFLOW_ID, TASK_ID = derive_demo_identifiers()

# Session-scoped, not fixed: multiple concurrent users (e.g. on a public
# multi-user deployment) each get their own subdirectory keyed by
# workflow_id, so a checkpoint DB, Memory store, diagnostics dir, or
# dashboard snapshot from one user's run can never collide with another's.
# Previously these were fixed paths shared by every caller - real, on a
# public deployment: two visitors clicking "Start Research" at the same
# time would have corrupted each other's runs. See _session_data_dir below
# and this module's docstring.
SESSIONS_ROOT = REPO_ROOT / "dashboard" / "data" / "sessions"


def _session_data_dir(workflow_id: str) -> Path:
    """Return (and create) one workflow's private data directory.

    workflow_id is produced by our own mandate builder or supplied via
    FULL_TEST_WORKFLOW_ID / the dashboard's session id, not arbitrary user
    input, so a light sanitization pass is enough here - same approach as
    FileBackedMemoryStore._path_for.
    """
    safe_name = "".join(
        ch if ch.isalnum() or ch in "-_." else "_" for ch in workflow_id
    )
    path = SESSIONS_ROOT / safe_name
    path.mkdir(parents=True, exist_ok=True)
    return path


# ---------------------------------------------------------------------------
# Offline data fixtures (real 120-ticker ETF data, no network)
# ---------------------------------------------------------------------------

def _normalize_offline_ohlc(
    *,
    open_price: float,
    high_price: float,
    low_price: float,
    close_price: float,
) -> tuple[float, float, float, float, bool]:
    """Repair only row-local OHLC bounds in the frozen demo fixture.

    Source rounding or formatting can leave a reported high or low inside
    another value from the same row. Canonicalizing the extrema keeps
    valid rows unchanged and does not interpolate, remove, or invent dates.
    The returned flag makes every adjusted row countable and disclosable.
    """

    values = (
        float(open_price),
        float(high_price),
        float(low_price),
        float(close_price),
    )
    normalized_high = max(values)
    normalized_low = min(values)
    adjusted = normalized_high != values[1] or normalized_low != values[2]
    return (
        values[0],
        normalized_high,
        normalized_low,
        values[3],
        adjusted,
    )


def _load_offline_panel() -> tuple[
    dict[str, tuple[PriceBar, ...]],
    dict[str, int],
]:
    from openpyxl import load_workbook

    wb = load_workbook(str(OFFLINE_PRICES_PATH), read_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    idx = {h: i for i, h in enumerate(header)}
    panel: dict[str, list[PriceBar]] = {}
    adjustments_by_symbol: dict[str, int] = {}
    for r in rows:
        ticker, dt, close = r[idx["ticker"]], r[idx["date"]], r[idx["close"]]
        if ticker is None or dt is None or close is None:
            continue
        open_price, high_price, low_price, close_price, adjusted = (
            _normalize_offline_ohlc(
                open_price=r[idx["open"]] or close,
                high_price=r[idx["high"]] or close,
                low_price=r[idx["low"]] or close,
                close_price=close,
            )
        )
        if adjusted:
            adjustments_by_symbol[ticker] = (
                adjustments_by_symbol.get(ticker, 0) + 1
            )
        panel.setdefault(ticker, []).append(PriceBar(
            symbol=ticker,
            timestamp=dt.replace(tzinfo=timezone.utc),
            open=open_price, high=high_price,
            low=low_price, close=close_price,
        ))
    return {k: tuple(v) for k, v in panel.items()}, adjustments_by_symbol


PANEL: dict[str, tuple[PriceBar, ...]] = {}
OFFLINE_OHLC_ADJUSTMENTS_BY_SYMBOL: dict[str, int] = {}
METADATA: dict[str, dict[str, Any]] = {}
OFFLINE_DATA_MAX_DATE = date.min


def _initialize_offline_data() -> None:
    """Load team fixtures once, after CLI parsing and dependency checks."""

    global PANEL
    global OFFLINE_OHLC_ADJUSTMENTS_BY_SYMBOL
    global METADATA
    global OFFLINE_DATA_MAX_DATE

    if PANEL:
        return
    _require_offline_data()
    PANEL, OFFLINE_OHLC_ADJUSTMENTS_BY_SYMBOL = _load_offline_panel()
    METADATA = _load_etf_metadata(OFFLINE_METADATA_PATH)
    if not PANEL:
        raise ValueError(
            f"No valid price series were found in {OFFLINE_PRICES_PATH}."
        )
    OFFLINE_DATA_MAX_DATE = max(
        bar.timestamp.date() for bars in PANEL.values() for bar in bars
    )


class OfflineDataService:
    """Serves PRICE_VOLUME + ETF_METADATA from the real fixtures - no network."""

    async def fetch(self, request: DataRequest) -> DataResponse:
        symbols = (
            request.asset_universe
            if isinstance(request.asset_universe, list) and request.asset_universe
            else list(PANEL)
        )
        now = datetime.now(timezone.utc)
        artifacts = []

        price_payload = {s: PANEL[s] for s in symbols if s in PANEL}
        if price_payload:
            adjustment_count = sum(
                OFFLINE_OHLC_ADJUSTMENTS_BY_SYMBOL.get(symbol, 0)
                for symbol in price_payload
            )
            artifacts.append(DataArtifact(
                artifact_id=f"{request.request_id}.prices",
                category=DataCategory.PRICE_VOLUME,
                description="Offline fixture prices (real ETF_historical_prices.xlsx)",
                data_reference="offline_fixture::prices",
                asset_scope=sorted(price_payload),
                provenance=[DataProvenance(
                    provenance_id=f"{request.request_id}.prices.prov", provider="offline_fixture",
                    source_reference=OFFLINE_PRICES_PATH.name, retrieved_at=now,
                    point_in_time_verified=False,
                )],
                analysis_payload=price_payload,
                limitations=(
                    [
                        f"Normalized high/low bounds on "
                        f"{adjustment_count} frozen-fixture rows "
                        "so each row's high and low bound its OHLC values; "
                        "dates and open/close values were unchanged."
                    ]
                    if adjustment_count
                    else []
                ),
            ))

        meta_payload = {
            s: {"category": METADATA[s]["category"], "fund_family": METADATA[s]["fund_family"],
                "issuer_tier": METADATA[s]["issuer_tier"]}
            for s in symbols if s in METADATA
        }
        if meta_payload:
            artifacts.append(DataArtifact(
                artifact_id=f"{request.request_id}.meta",
                category=DataCategory.ETF_METADATA,
                description="Offline fixture ETF metadata (real ETF_info.xlsx)",
                data_reference="offline_fixture::etf_metadata",
                asset_scope=sorted(meta_payload),
                provenance=[DataProvenance(
                    provenance_id=f"{request.request_id}.meta.prov", provider="offline_fixture",
                    source_reference=OFFLINE_METADATA_PATH.name, retrieved_at=now,
                    point_in_time_verified=False,
                )],
                analysis_payload=meta_payload,
            ))

        return DataResponse(
            response_id=f"{request.request_id}.response", request_id=request.request_id,
            lineage=request.lineage, as_of_date=request.as_of_date,
            complete=True, artifacts=artifacts,
        )


class OfflineBacktestResolver:
    """Resolves backtest bars from the real fixture panel - no network.

    Handles single-ticker candidates, Quant pairs, Fundamental benchmark
    lists, and Technical multi-asset sleeves without changing any agent
    contract.
    """

    async def resolve(self, request: BacktestRequest) -> ResolvedBacktestData:
        params = request.candidate.parameters
        symbols: list[str] = []
        for key in ("ticker_a", "ticker_b", "symbol", "ticker"):
            value = params.get(key)
            if value:
                symbols.append(str(value))
        symbols.extend(str(s) for s in params.get("benchmark_tickers", []))
        raw_sleeves = params.get("sleeves", [])
        if isinstance(raw_sleeves, list):
            symbols.extend(
                str(sleeve["symbol"])
                for sleeve in raw_sleeves
                if isinstance(sleeve, Mapping) and sleeve.get("symbol")
            )
        benchmark = getattr(getattr(request, "plan", None), "benchmark", None)
        if benchmark:
            symbols.append(str(benchmark))
        bars = tuple(bar for s in dict.fromkeys(symbols) if s in PANEL for bar in PANEL[s])
        return ResolvedBacktestData(data_references=("offline_fixture",), bars=bars)


# ---------------------------------------------------------------------------
# Stub Technical Trader node (see module docstring for why)
# ---------------------------------------------------------------------------

def _stub_technical_trader_package(
    mandate: PMMandate,
    round_number: int,
) -> TraderStrategyPackage:
    """A schema-valid stand-in used only without provider configuration.

    Fills the graph's technical_trader slot so the full topology (3 trader
    branches -> join -> Risk -> Reporting -> PM) can be exercised end-to-end.
    Values are illustrative and are not Technical Trader research output.
    """
    from protocols import (
        CandidateRuleSpecification,
        ConstraintCheckStatus,
        MandateConstraintAssessment,
    )

    lineage = TaskLineage(
        workflow_id=mandate.workflow_id,
        task_id=f"{mandate.task_id}.round-{round_number}.technical.trader",
        parent_task_id=mandate.task_id,
        source_task_id=mandate.task_id,
        attempt=1,
    )
    candidate = CandidateRuleSpecification(
        strategy_name="[STUB] SMA crossover on QQQ",
        hypothesis="[STUB - not a real model call] Illustrative placeholder only.",
        rule_summary="[STUB] 20/50-day SMA crossover, long-only.",
        executor_id="technical_trader.stub_demo.v0",
        asset_eligibility_logic="[STUB]",
        signal_logic="[STUB]",
        position_logic="[STUB]",
        entry_logic="[STUB]",
        exit_logic="[STUB]",
        rebalancing_logic="[STUB]",
        parameters={"ticker": "QQQ", "short_window": 20, "long_window": 50},
        specialty_evidence_ids=["technical_trader.stub_demo"],
        specialty_evidence_usage={"technical_trader.stub_demo": "[STUB placeholder evidence]"},
        candidate_id=f"{lineage.task_id}.candidate",
        trader_id=SpecialistId.TECHNICAL_TRADER,
        lineage=lineage.child("candidate"),
    )
    return TraderStrategyPackage(
        package_id=f"{lineage.task_id}.package",
        candidate_id=candidate.candidate_id,
        trader_id=SpecialistId.TECHNICAL_TRADER,
        lineage=lineage,
        mandate_reference=mandate.reference(),
        status=RunStatus.FAILED,  # honestly not eligible - it's a stub, not a real result
        candidate_rule=candidate,
        constraint_assessment=MandateConstraintAssessment(
            status=ConstraintCheckStatus.NOT_EVALUATED,
            requires_risk_validation=False,
        ),
        failures=[{
            "stage": "technical_trader.stub_demo",
            "message": (
                "This demo run stubs Technical Trader because no ModelClient "
                "provider was configured. It is not eligible for Risk review; "
                "set the Technical Trader environment variables described in "
                ".env.example to run the real agent."
            ),
            "retryable": False,
        }],
        eligible_for_risk_review=False,
    )


# ---------------------------------------------------------------------------
# PM intake (Memory read/write are now real - see main(); PM decision is
# no longer scripted here at all - omitting it from ProductionNodeSet makes
# the graph default to its own real human_pm_decision_node, a durable
# LangGraph interrupt. See this module's docstring for why that requires a
# persistent (SQLite) checkpointer instead of the in-process MemorySaver
# this script used previously.)
# ---------------------------------------------------------------------------

def pm_intake_node(state: Mapping[str, Any]) -> Mapping[str, Any]:
    del state
    return {}


class RecordingReportingNode:
    def __init__(self, reporting_agent: ReportingAgentImpl) -> None:
        self._agent = reporting_agent
        self.requests: list[ReportingRequest] = []

    async def __call__(self, state: Mapping[str, Any]) -> Mapping[str, Any]:
        request = ReportingRequest.model_validate(state.get("reporting_request"))
        self.requests.append(request)
        output = await self._agent.report(request)
        return {"reporting_output": output.model_dump(mode="json")}


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def _load_mandate_and_payload(mandate_path: Path | None) -> tuple[PMMandate, dict[str, Any]]:
    """Load a dashboard-created mandate, or use the documented demo default.

    Returns both the validated mandate and the raw payload dict, since the
    payload may also carry ``active_specialists`` (staffing) for this run -
    a sibling key to ``pm_mandate``, not part of the mandate itself.
    """

    if mandate_path is None:
        mandate = PMMandate(
            workflow_id=WORKFLOW_ID,
            task_id=TASK_ID,
            as_of_date=OFFLINE_DATA_MAX_DATE,
            investment_objective=(
                "Full research-loop integration demo across the 120-ETF universe."
            ),
            permitted_asset_universe=list(PANEL),
        )
        return mandate, {}

    payload = json.loads(mandate_path.read_text(encoding="utf-8"))
    raw_mandate = payload.get("pm_mandate", payload)
    mandate = PMMandate.model_validate(raw_mandate)
    if mandate.as_of_date > OFFLINE_DATA_MAX_DATE:
        raise ValueError(
            "The offline ETF fixture ends on "
            f"{OFFLINE_DATA_MAX_DATE.isoformat()}, but the PM mandate requested "
            f"{mandate.as_of_date.isoformat()}. Choose an available as-of date."
        )
    return mandate, payload


def _build_nodes(
    memory_store: FileBackedMemoryStore, technical_diagnostics_path: Path,
) -> "ProductionNodeSet":
    """Build the ProductionNodeSet shared by both start and resume paths."""

    async def memory_read_node(state: Mapping[str, Any]) -> dict[str, Any]:
        # memory_read runs first in the graph, before the "prepare" node
        # that derives a top-level "workflow_id" key from the mandate - so
        # fall back to the mandate itself, which is always present.
        workflow_id = str(state.get("workflow_id") or state["pm_mandate"]["workflow_id"])
        context = await memory_store.load_context(workflow_id)
        return {
            "memory_context": context.model_dump(mode="json"),
            "round_audit_summary_reference": f"{workflow_id}.round-{state.get('round_number', 1)}.audit",
            "round_history_reference": f"{workflow_id}.history",
        }

    async def memory_write_node(state: Mapping[str, Any]) -> dict[str, Any]:
        pm_decision = PMDecision.model_validate(state["pm_decision"])
        reporting_output = state.get("reporting_output") or {}
        risk_response = state.get("risk_review_response") or {}
        mandate_lessons = state.get("pm_mandate", {}).get("prior_round_lessons", [])
        pivot_lessons = (
            [str(lesson) for lesson in mandate_lessons]
            if isinstance(mandate_lessons, list)
            else []
        )
        lessons = list(dict.fromkeys(
            ([pm_decision.rationale] if pm_decision.rationale else [])
            + pivot_lessons
        ))
        record = MemoryRecord(
            record_id=f"{state['workflow_id']}.round-{state.get('round_number', 1)}.record",
            workflow_id=str(state.get("workflow_id") or state["pm_mandate"]["workflow_id"]),
            mandate_task_id=str(state["pm_mandate"]["task_id"]),
            result_references=list(reporting_output.get("surviving_candidate_ids") or []),
            critiques=[
                str(c) for c in (risk_response.get("critiques") or [])
            ] if isinstance(risk_response.get("critiques"), list) else [],
            pm_decision=pm_decision,
            lessons_for_future_rounds=lessons,
        )
        record_id = await memory_store.record(record)
        return {"memory_record_id": record_id}

    # --- Real Fundamental Trader ---
    fundamental_agent = FundamentalTraderAgent(
        data_service=OfflineDataService(),
        backtest_engine=DeterministicBacktestEngine(
            data_resolver=OfflineBacktestResolver(),
            strategy_executors=[category_deviation_executor],
        ),
        validation_split_policy=PercentileValidationSplitPolicy(train_fraction=0.9),
    )
    fundamental_runtime = FundamentalTraderRuntime(agent=fundamental_agent)

    async def fundamental_trader_node(state: Mapping[str, Any]) -> dict[str, Any]:
        current_mandate = PMMandate.model_validate(state["pm_mandate"])
        package = await fundamental_runtime.research(current_mandate, execution_context={
            "run_id": f"{current_mandate.workflow_id}.run",
            "round_number": state.get("round_number", 1), "attempt": 1,
        })
        return {"fundamental_trader_package": package.model_dump(mode="json")}

    # --- Real Quant Trader (Shaurya's) ---
    quant_agent = QuantTraderAgent(
        data_service=OfflineDataService(),
        backtest_engine=DeterministicBacktestEngine(
            data_resolver=OfflineBacktestResolver(),
            strategy_executors=[cross_asset_spread_executor],
        ),
        validation_split_policy=PercentileValidationSplitPolicy(train_fraction=0.9),
    )
    quant_runtime = QuantTraderRuntime(agent=quant_agent)

    async def quant_trader_node(state: Mapping[str, Any]) -> dict[str, Any]:
        current_mandate = PMMandate.model_validate(state["pm_mandate"])
        package = await quant_runtime.research(current_mandate, execution_context={
            "run_id": f"{current_mandate.workflow_id}.run",
            "round_number": state.get("round_number", 1), "attempt": 1,
        })
        return {"quant_trader_package": package.model_dump(mode="json")}

    # --- Technical Trader: real if credentials are configured, else stubbed ---
    try:
        technical_execution_policy = TechnicalExecutionPolicy()
        technical_model_client = create_technical_model_client_from_env(
            execution_policy=technical_execution_policy,
        )
        technical_backtest_engine = DeterministicBacktestEngine(
            data_resolver=OfflineBacktestResolver(),
            strategy_executors=list(TECHNICAL_STRATEGY_EXECUTORS),
        )
        technical_runtime = create_technical_trader_runtime(
            model_client=technical_model_client,
            data_service=OfflineDataService(),
            backtest_engine=technical_backtest_engine,
            available_executors=[e.executor_id for e in TECHNICAL_STRATEGY_EXECUTORS],
            validation_split_policy=HorizonMatchedValidationSplitPolicy(),
            benchmark_symbol=TECHNICAL_TRADER_BENCHMARK_SYMBOL,
            diagnostics_sink=JsonFileTechnicalDiagnosticsSink(
                technical_diagnostics_path
            ),
            execution_policy=technical_execution_policy,
        )
        print(
            "Technical Trader: REAL - "
            f"{technical_model_client.__class__.__name__} configured from environment."
        )

        async def technical_trader_node(state: Mapping[str, Any]) -> dict[str, Any]:
            current_mandate = PMMandate.model_validate(state["pm_mandate"])
            package = await technical_runtime.research(current_mandate, execution_context={
                "run_id": f"{current_mandate.workflow_id}.run",
                "round_number": state.get("round_number", 1), "attempt": 1,
            })
            return {"technical_trader_package": package.model_dump(mode="json")}

    except TechnicalModelConfigurationError as config_error:
        if os.environ.get("TECHNICAL_TRADER_MODEL_PROVIDER", "").strip():
            raise RuntimeError(
                "Technical Trader provider configuration is invalid. Review "
                ".env.example and the underlying error: "
                f"{config_error}"
            ) from config_error
        print(
            "Technical Trader: STUBBED - no model provider configured "
            f"({config_error}). Set TECHNICAL_TRADER_MODEL_PROVIDER, "
            "TECHNICAL_TRADER_MODEL, and the matching API key to use the "
            "real agent - see agents/technical_trader/docs/integration.md."
        )

        def technical_trader_node(state: Mapping[str, Any]) -> dict[str, Any]:
            current_mandate = PMMandate.model_validate(state["pm_mandate"])
            package = _stub_technical_trader_package(
                current_mandate,
                int(state.get("round_number", 1)),
            )
            return {"technical_trader_package": package.model_dump(mode="json")}

    # --- Real Risk + Reporting ---
    risk_agent = RiskAgentImpl()
    # Reporting can write a Gemini memo when the runner is launched in an
    # environment that has Emma's GEMINI_API_KEY.  The deterministic
    # comparison remains available without credentials, so classmates can
    # still run the complete offline loop locally.
    try:
        from services.gemini_model_client import GeminiModelClient

        reporting_agent = ReportingAgentImpl(model_client=GeminiModelClient())
        print("Reporting Agent: Gemini memo generation enabled.")
    except (ImportError, KeyError):
        reporting_agent = ReportingAgentImpl()
        print(
            "Reporting Agent: structured comparison only. Set GEMINI_API_KEY "
            "to enable the optional narrative memo."
        )
    reporting = RecordingReportingNode(reporting_agent)

    return ProductionNodeSet(
        memory_read=memory_read_node,
        pm_intake=pm_intake_node,
        technical_trader=technical_trader_node,
        fundamental_trader=fundamental_trader_node,
        quant_trader=quant_trader_node,
        risk_review=make_risk_review_node(risk_agent),
        reporting=reporting,
        memory_write=memory_write_node,
        # pm_decision intentionally omitted: the graph defaults to its own
        # real human_pm_decision_node (a durable LangGraph interrupt), which
        # is what makes an actual live PM decision - via the dashboard,
        # resumed by this script's --resume mode - possible instead of a
        # scripted stand-in.
    )


def _print_state_summary(final_state: Mapping[str, Any]) -> None:
    print("=== Trader packages ===")
    for key in ("technical_trader_package", "fundamental_trader_package", "quant_trader_package"):
        pkg = final_state.get(key)
        if pkg:
            print(f"{key}: status={pkg['status']} eligible_for_risk_review={pkg['eligible_for_risk_review']}")

    print()
    print("=== Risk review ===")
    risk_response = final_state.get("risk_review_response")
    if risk_response:
        print(f"approved candidates: {risk_response.get('approved_candidate_ids', risk_response)}")

    print()
    print("=== Reporting output ===")
    reporting_output = final_state.get("reporting_output")
    if reporting_output:
        print(f"surviving_candidate_ids: {reporting_output.get('surviving_candidate_ids')}")

    print()
    if final_state.get("pending_human_action"):
        print("=== Awaiting PM decision ===")
        print(
            "The graph paused at a durable interrupt for round "
            f"{final_state.get('round_number', 1)}. Resume with:\n"
            "  python scripts/run_full_research_loop_demo.py --resume "
            f"--run-id {final_state.get('workflow_id')} --decision-json <path>"
        )
        return

    print("=== PM decision ===")
    pm_decision = final_state.get("pm_decision")
    if pm_decision:
        print(f"decision: {pm_decision.get('decision')}, rationale: {pm_decision.get('rationale')}")

    print()
    print(f"memory_record_id: {final_state.get('memory_record_id')}")


async def main(
    mandate_path: Path | None = None,
    *,
    resume: bool = False,
    run_id: str | None = None,
    decision_path: Path | None = None,
) -> None:
    if langgraph_import_error is not None:
        raise RuntimeError(
            "Full-loop dependencies are not installed. Run "
            "pip install -e '.[full-demo]'."
        ) from langgraph_import_error

    _initialize_offline_data()

    # Determine the session-scoped workflow_id *before* opening any
    # per-session file (checkpoint DB, Memory store, diagnostics, snapshot)
    # - this is what lets multiple concurrent dashboard users run real,
    # independent workflows without colliding on a shared file.
    mandate: PMMandate | None = None
    payload: dict[str, Any] = {}
    if resume:
        if not run_id or not decision_path:
            print("--resume requires both --run-id and --decision-json")
            return
        session_workflow_id = run_id
    else:
        mandate, payload = _load_mandate_and_payload(mandate_path)
        session_workflow_id = mandate.workflow_id

    session_dir = _session_data_dir(session_workflow_id)
    checkpoint_db_path = session_dir / "workflow_checkpoints.sqlite"
    technical_diagnostics_path = session_dir / "technical_trader_diagnostics"
    memory_store_dir = session_dir / "memory"
    snapshot_path = session_dir / "workflow_snapshot.json"

    memory_store = FileBackedMemoryStore(memory_store_dir)

    async with AsyncSqliteSaver.from_conn_string(str(checkpoint_db_path)) as checkpointer:
        nodes = _build_nodes(memory_store, technical_diagnostics_path)
        # Keep the workflow aligned with Risk's three-round validation-touch
        # budget. The PM can request another round through round two; round
        # three requires a select or reject decision.
        compiled = compile_production_workflow(nodes, checkpointer=checkpointer, max_rounds=3)
        runner = WorkflowRunner(
            compiled_graph=compiled,
            snapshot_writer=lambda state: write_dashboard_snapshot(state, snapshot_path),
        )

        if resume:
            decision_payload = json.loads(decision_path.read_text(encoding="utf-8"))
            pm_decision = decision_payload.get("pm_decision", decision_payload)
            state_update = decision_payload.get("state_update")
            print(f"Resuming workflow {run_id} with decision "
                  f"{pm_decision.get('decision')!r}...")
            final_state = await runner.resume_workflow(
                run_id, pm_decision, state_update=state_update,
            )
        else:
            workflow_input: dict[str, Any] = {
                "pm_mandate": mandate.model_dump(mode="json"),
                "run_id": mandate.workflow_id,
            }
            if payload.get("active_specialists"):
                workflow_input["active_specialists"] = payload["active_specialists"]

            config = {"configurable": {"thread_id": mandate.workflow_id}}
            existing_checkpoint = await compiled.aget_state(config)
            existing_values = getattr(
                existing_checkpoint,
                "values",
                existing_checkpoint,
            )
            if isinstance(existing_values, Mapping) and existing_values:
                raise RuntimeError(
                    "A checkpoint already exists for workflow_id "
                    f"'{mandate.workflow_id}'. Start with a fresh "
                    "FULL_TEST_WORKFLOW_ID or use --resume with that ID."
                )

            print("Running one research-loop round (or resuming to the next PM decision)...")
            print("  Real: Fundamental Trader, Quant Trader, Risk Agent, Reporting Agent, Memory")
            print("  Real (durable interrupt): PM decision")
            print()
            final_state = await runner.start_workflow(workflow_input, publish_progress=True)

        print(f"Dashboard snapshot written to {snapshot_path}")
        _print_state_summary(final_state)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Run or resume the offline research loop, optionally driven by dashboard JSON files."
    )
    parser.add_argument(
        "--mandate-json",
        type=Path,
        help="Path to a PMMandate JSON file or a WorkflowInput JSON file (start mode only).",
    )
    parser.add_argument(
        "--resume",
        action="store_true",
        help="Resume a paused PM-decision interrupt instead of starting a new run.",
    )
    parser.add_argument(
        "--run-id",
        type=str,
        help="workflow_id of the run to resume (required with --resume).",
    )
    parser.add_argument(
        "--decision-json",
        type=Path,
        help=(
            "Path to a JSON file with {\"pm_decision\": {...}, "
            "\"state_update\": {...}} (required with --resume)."
        ),
    )
    args = parser.parse_args()
    asyncio.run(main(
        args.mandate_json,
        resume=args.resume,
        run_id=args.run_id,
        decision_path=args.decision_json,
    ))
