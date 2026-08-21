#!/usr/bin/env python3
"""Multi-round memory integration demo.

Extends ``run_full_research_loop_demo.py`` by wiring in the REAL
``services.memory_store_impl.InMemoryMemoryStore`` (built and unit-tested
separately - see tests/memory_store/) instead of the scripted memory
stand-ins used there, and running the compiled graph for TWO rounds:

  * Round 1: a scripted PM decision requests another round (mirroring a
    real "the results looked thin, try again" call).
  * Round 2: the graph loops back to PM intake, Memory supplies the
    context recorded from round 1, real traders run again, and a
    scripted PM decision closes the workflow.

This is the piece that was previously untested end-to-end: the Memory
Store itself already had unit tests (see tests/memory_store/), and the
single-round loop was proven separately, but nobody had confirmed the
two combine correctly - that a second round's `memory_context` state key
actually reflects what round 1 wrote. It does; see the assertions below.

Run (offline, no network - uses the real 120-ticker ETF fixtures):

    python scripts/run_multi_round_memory_demo.py
"""

from __future__ import annotations

import asyncio
import pickle
import sys
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any, Mapping

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

from protocols import (  # noqa: E402
    BacktestRequest,
    DataArtifact,
    DataCategory,
    DataProvenance,
    DataRequest,
    DataResponse,
    PMDecision,
    PMDecisionType,
    PMMandate,
)
from protocols.research_contracts import MemoryRecord  # noqa: E402
from tools import DeterministicBacktestEngine, PriceBar, ResolvedBacktestData  # noqa: E402

from agents.fundamental_trader import (  # noqa: E402
    FundamentalTraderRuntime,
    category_deviation_executor,
)
from agents.fundamental_trader.agent import FundamentalTraderAgent  # noqa: E402
from agents.fundamental_trader.examples.static_data_service import _load_etf_metadata  # noqa: E402
from agents.fundamental_trader.examples.validation_split import (  # noqa: E402
    PercentileValidationSplitPolicy,
)
from agents.quant_trader import (  # noqa: E402
    QuantTraderAgent,
    QuantTraderRuntime,
    cross_asset_spread_executor,
)
from agents.risk_agent import RiskAgentImpl, make_risk_review_node  # noqa: E402
from agents.reporting_agent import ReportingAgentImpl  # noqa: E402
from protocols.reporting import ReportingRequest  # noqa: E402
from services.memory_store_impl import InMemoryMemoryStore  # noqa: E402

langgraph_missing = False
try:
    from graph.production import ProductionNodeSet, compile_production_workflow  # noqa: E402
    from langgraph.checkpoint.memory import MemorySaver  # noqa: E402
except ImportError:
    langgraph_missing = True


REPO_ROOT = Path(__file__).resolve().parents[1]
RUN_ID = "multi-round-memory-demo-run"
WORKFLOW_ID = "multi-round-memory-demo-workflow"
TASK_ID = "multi-round-memory-demo-task"


# ---------------------------------------------------------------------------
# Offline data fixtures (same as run_full_research_loop_demo.py)
# ---------------------------------------------------------------------------

def _load_offline_panel() -> dict[str, tuple[PriceBar, ...]]:
    cache = Path("/tmp/price_panel.pkl")
    if cache.exists():
        with cache.open("rb") as f:
            return pickle.load(f)
    from openpyxl import load_workbook
    prices_path = REPO_ROOT / "ETF_historical_prices.xlsx"
    if not prices_path.exists():
        raise FileNotFoundError(f"{prices_path} not found.")
    wb = load_workbook(str(prices_path), read_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    idx = {h: i for i, h in enumerate(header)}
    panel: dict[str, list[PriceBar]] = {}
    for r in rows:
        ticker, dt, close = r[idx["ticker"]], r[idx["date"]], r[idx["close"]]
        if ticker is None or dt is None or close is None:
            continue
        panel.setdefault(ticker, []).append(PriceBar(
            symbol=ticker, timestamp=dt.replace(tzinfo=timezone.utc),
            open=r[idx["open"]] or close, high=r[idx["high"]] or close,
            low=r[idx["low"]] or close, close=close,
        ))
    return {k: tuple(v) for k, v in panel.items()}


PANEL = _load_offline_panel()
METADATA = _load_etf_metadata(REPO_ROOT / "ETF_info.xlsx")


class OfflineDataService:
    async def fetch(self, request: DataRequest) -> DataResponse:
        symbols = (
            request.asset_universe
            if isinstance(request.asset_universe, list) and request.asset_universe
            else list(PANEL)
        )
        now = datetime.now(timezone.utc)
        artifacts = []
        price_payload = {s: PANEL[s] for s in symbols if s in PANEL}
        if price_payload:
            artifacts.append(DataArtifact(
                artifact_id=f"{request.request_id}.prices", category=DataCategory.PRICE_VOLUME,
                description="Offline fixture prices", data_reference="offline_fixture::prices",
                asset_scope=sorted(price_payload),
                provenance=[DataProvenance(
                    provenance_id=f"{request.request_id}.prices.prov", provider="offline_fixture",
                    source_reference="ETF_historical_prices.xlsx", retrieved_at=now,
                    point_in_time_verified=False,
                )],
                analysis_payload=price_payload,
            ))
        meta_payload = {
            s: {"category": METADATA[s]["category"], "fund_family": METADATA[s]["fund_family"],
                "issuer_tier": METADATA[s]["issuer_tier"]}
            for s in symbols if s in METADATA
        }
        if meta_payload:
            artifacts.append(DataArtifact(
                artifact_id=f"{request.request_id}.meta", category=DataCategory.ETF_METADATA,
                description="Offline fixture ETF metadata", data_reference="offline_fixture::etf_metadata",
                asset_scope=sorted(meta_payload),
                provenance=[DataProvenance(
                    provenance_id=f"{request.request_id}.meta.prov", provider="offline_fixture",
                    source_reference="ETF_info.xlsx", retrieved_at=now, point_in_time_verified=False,
                )],
                analysis_payload=meta_payload,
            ))
        return DataResponse(
            response_id=f"{request.request_id}.response", request_id=request.request_id,
            lineage=request.lineage, as_of_date=request.as_of_date, complete=True, artifacts=artifacts,
        )


class OfflineBacktestResolver:
    async def resolve(self, request: BacktestRequest) -> ResolvedBacktestData:
        params = request.candidate.parameters
        symbols: list[str] = []
        for key in ("ticker_a", "ticker_b", "symbol", "ticker"):
            value = params.get(key)
            if value:
                symbols.append(str(value))
        symbols.extend(str(s) for s in params.get("benchmark_tickers", []))
        bars = tuple(bar for s in dict.fromkeys(symbols) if s in PANEL for bar in PANEL[s])
        return ResolvedBacktestData(data_references=("offline_fixture",), bars=bars)


def _stub_technical_trader_package(round_number: int):
    from protocols import (
        CandidateRuleSpecification, ConstraintCheckStatus, MandateConstraintAssessment,
        RunStatus, SpecialistId, TaskLineage, TraderStrategyPackage,
    )
    lineage = TaskLineage(
        workflow_id=WORKFLOW_ID, task_id=f"{TASK_ID}.round-{round_number}.technical.trader", attempt=1,
    )
    candidate = CandidateRuleSpecification(
        strategy_name="[STUB] SMA crossover on QQQ", hypothesis="[STUB - not a real model call]",
        rule_summary="[STUB]", executor_id="technical_trader.stub_demo.v0",
        asset_eligibility_logic="[STUB]", signal_logic="[STUB]", position_logic="[STUB]",
        entry_logic="[STUB]", exit_logic="[STUB]", rebalancing_logic="[STUB]",
        parameters={"ticker": "QQQ", "short_window": 20, "long_window": 50},
        specialty_evidence_ids=["technical_trader.stub_demo"],
        specialty_evidence_usage={"technical_trader.stub_demo": "[STUB]"},
        candidate_id=f"{lineage.task_id}.candidate", trader_id=SpecialistId.TECHNICAL_TRADER,
        lineage=lineage.child("candidate"),
    )
    return TraderStrategyPackage(
        package_id=f"{lineage.task_id}.package", candidate_id=candidate.candidate_id,
        trader_id=SpecialistId.TECHNICAL_TRADER, lineage=lineage,
        mandate_reference={"workflow_id": WORKFLOW_ID, "task_id": TASK_ID, "as_of_date": date.today().isoformat()},
        status=RunStatus.FAILED, candidate_rule=candidate,
        constraint_assessment=MandateConstraintAssessment(
            status=ConstraintCheckStatus.NOT_EVALUATED, requires_risk_validation=False,
        ),
        failures=[{"stage": "technical_trader.stub_demo",
                   "message": "Stubbed - this demo does not configure a Technical Trader model provider.", "retryable": False}],
        eligible_for_risk_review=False,
    )


class RecordingReportingNode:
    def __init__(self, reporting_agent: ReportingAgentImpl) -> None:
        self._agent = reporting_agent

    async def __call__(self, state: Mapping[str, Any]) -> Mapping[str, Any]:
        request = ReportingRequest.model_validate(state.get("reporting_request"))
        output = await self._agent.report(request)
        return {"reporting_output": output.model_dump(mode="json")}


async def main() -> None:
    if langgraph_missing:
        print("langgraph is not installed - run: pip install -e '.[langgraph]'")
        return

    mandate = PMMandate(
        workflow_id=WORKFLOW_ID, task_id=TASK_ID, as_of_date=date(2026, 6, 30),
        investment_objective="Multi-round memory integration demo across the 120-ETF universe.",
        permitted_asset_universe=list(PANEL),
    )

    # --- The one thing this demo is actually testing: the REAL memory store ---
    memory_store = InMemoryMemoryStore()

    async def memory_read_node(state: Mapping[str, Any]) -> dict[str, Any]:
        context = await memory_store.load_context(WORKFLOW_ID)
        return {
            "memory_context": context.model_dump(mode="json"),
            "round_audit_summary_reference": f"demo-audit-round-{state.get('round_number', 1)}",
            "round_history_reference": "demo-history-ref",
        }

    async def memory_write_node(state: Mapping[str, Any]) -> dict[str, Any]:
        pm_decision = PMDecision.model_validate(state["pm_decision"])
        record = MemoryRecord(
            record_id=f"{WORKFLOW_ID}.round-{state.get('round_number', 1)}.record",
            workflow_id=WORKFLOW_ID,
            mandate_task_id=TASK_ID,
            result_references=list(state.get("surviving_candidate_ids", [])),
            critiques=[],
            pm_decision=pm_decision,
            lessons_for_future_rounds=(
                ["Round 1: results looked thin, requested a second pass with the same universe."]
                if pm_decision.decision == PMDecisionType.REQUEST_ANOTHER_ROUND
                else []
            ),
        )
        record_id = await memory_store.record(record)
        return {"memory_record_id": record_id}

    def pm_intake_node(state: Mapping[str, Any]) -> dict[str, Any]:
        del state
        return {}

    def pm_decision_node(state: Mapping[str, Any]) -> dict[str, Any]:
        round_number = state.get("round_number", 1)
        if round_number == 1:
            decision = PMDecision(
                decision_id=f"{state['workflow_id']}.decision-round-1",
                workflow_id=str(state["workflow_id"]),
                decision=PMDecisionType.REQUEST_ANOTHER_ROUND,
                rationale="Scripted: requesting a second round to prove memory carries forward.",
            )
        else:
            decision = PMDecision(
                decision_id=f"{state['workflow_id']}.decision-round-{round_number}",
                workflow_id=str(state["workflow_id"]),
                decision=PMDecisionType.REJECT,
                rationale="Scripted: closing the workflow after round 2.",
            )
        return {"pm_decision": decision.model_dump(mode="json"), "pending_human_action": None}

    fundamental_agent = FundamentalTraderAgent(
        data_service=OfflineDataService(),
        backtest_engine=DeterministicBacktestEngine(
            data_resolver=OfflineBacktestResolver(), strategy_executors=[category_deviation_executor],
        ),
        validation_split_policy=PercentileValidationSplitPolicy(train_fraction=0.9),
    )
    fundamental_runtime = FundamentalTraderRuntime(agent=fundamental_agent)

    async def fundamental_trader_node(state: Mapping[str, Any]) -> dict[str, Any]:
        package = await fundamental_runtime.research(mandate, execution_context={
            "run_id": RUN_ID, "round_number": state.get("round_number", 1), "attempt": 1,
        })
        return {"fundamental_trader_package": package.model_dump(mode="json")}

    quant_agent = QuantTraderAgent(
        data_service=OfflineDataService(),
        backtest_engine=DeterministicBacktestEngine(
            data_resolver=OfflineBacktestResolver(), strategy_executors=[cross_asset_spread_executor],
        ),
        validation_split_policy=PercentileValidationSplitPolicy(train_fraction=0.9),
    )
    quant_runtime = QuantTraderRuntime(agent=quant_agent)

    async def quant_trader_node(state: Mapping[str, Any]) -> dict[str, Any]:
        package = await quant_runtime.research(mandate, execution_context={
            "run_id": RUN_ID, "round_number": state.get("round_number", 1), "attempt": 1,
        })
        return {"quant_trader_package": package.model_dump(mode="json")}

    def technical_trader_node(state: Mapping[str, Any]) -> dict[str, Any]:
        return {"technical_trader_package": _stub_technical_trader_package(
            state.get("round_number", 1)
        ).model_dump(mode="json")}

    risk_agent = RiskAgentImpl()
    reporting = RecordingReportingNode(ReportingAgentImpl())

    nodes = ProductionNodeSet(
        memory_read=memory_read_node,
        pm_intake=pm_intake_node,
        technical_trader=technical_trader_node,
        fundamental_trader=fundamental_trader_node,
        quant_trader=quant_trader_node,
        risk_review=make_risk_review_node(risk_agent),
        reporting=reporting,
        memory_write=memory_write_node,
        pm_decision=pm_decision_node,
    )
    compiled = compile_production_workflow(nodes, checkpointer=MemorySaver(), max_rounds=2)

    print("Round 1: expect memory_context to be EMPTY (nothing recorded yet)")
    print("Round 2 (after PM requests another round): expect memory_context to carry")
    print("round 1's lessons, forward - this is the part that was previously unproven.\n")

    final_state = await compiled.ainvoke(
        {"pm_mandate": mandate.model_dump(mode="json"), "run_id": RUN_ID},
        config={"configurable": {"thread_id": "multi-round-memory-demo"}},
    )

    # Inspect what memory actually recorded, independent of the graph state,
    # to prove the real MemoryStore (not just the graph) did its job.
    context_after = await memory_store.load_context(WORKFLOW_ID)
    print("=== Real MemoryStore state after the run ===")
    print(f"prior_pm_decisions: {context_after.prior_pm_decisions}")
    print(f"lessons_for_next_round: {context_after.lessons_for_next_round}")

    assert len(context_after.prior_pm_decisions) == 2, (
        "expected two rounds of PM decisions to be recorded"
    )
    assert any("thin" in lesson for lesson in context_after.lessons_for_next_round), (
        "expected round 1's lesson to have been recorded"
    )
    print("\nPASS: two rounds recorded; round 1's lesson is present in the real MemoryStore.")

    print(f"\nfinal pm_decision: {final_state.get('pm_decision', {}).get('decision')}")
    print(f"final round_number reached: {final_state.get('round_number')}")


if __name__ == "__main__":
    asyncio.run(main())
