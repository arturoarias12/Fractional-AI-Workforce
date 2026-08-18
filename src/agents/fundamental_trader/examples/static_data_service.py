"""Composite DataService for the standalone Fundamental Trader demo.

The shared ``services.data_service.YFinanceDataService`` (Yiran's
workstream) only builds ``PRICE_VOLUME`` artifacts today - verified by
reading its source, not assumed. Fundamental Trader also needs
``ETF_METADATA`` (category, fund family) to run its ISSUER_SCALE_TIER
heuristic, which nothing shared serves yet. ``FundamentalMetadataDataService``
below is the seam that will shrink or disappear once the shared DataService
grows an ETF_METADATA path: it delegates PRICE_VOLUME to the real shared
service unchanged, and adds ETF_METADATA itself by reading ``ETF_info.xlsx``
directly (requires ``pip install -e .[fundamental-demo]``; the workbook
itself is a local fixture, not part of the package - place a copy at the
repo root, or pass an explicit path).

ETF_info.xlsx was inspected directly for this project: ``marketCap``,
``sector``, and ``industry`` are null/empty for effectively the whole
universe. Only ``category`` and ``fundFamily`` are populated, which is why
``rule_generator.classify_issuer_tier`` exists - it turns ``fundFamily``
into the ISSUER_SCALE_TIER major/boutique heuristic used in place of the
fundamentals that aren't available. See ``docs/fundamental_trader.md``.
"""

from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

from protocols import (
    DataArtifact,
    DataCategory,
    DataProvenance,
    DataRequest,
    DataResponse,
)

from services import YFinanceDataService

from ..rule_generator import classify_issuer_tier

# Reuses the same 120-ticker fixture universe already established for Quant
# Trader's demo, rather than maintaining a second copy.
from agents.quant_trader.examples.static_data_service import DEFAULT_UNIVERSE

DEFAULT_ETF_INFO_PATH = Path("ETF_info.xlsx")

_METADATA_LIMITATIONS = [
    "ETF_info.xlsx does not populate marketCap, sector, or industry for this "
    "universe (verified by inspection). Only category and fundFamily are "
    "used; ISSUER_SCALE_TIER (major/boutique) is a heuristic proxy built "
    "from fundFamily, not a licensed fundamental data field.",
]


def _load_etf_metadata(xlsx_path: Path) -> dict[str, dict[str, str]]:
    """Return ``{ticker: {"category", "fund_family", "issuer_tier"}}``."""
    wb = load_workbook(str(xlsx_path), read_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    idx = {name: i for i, name in enumerate(header)}

    required = ("ticker", "category", "fundFamily")
    missing = [name for name in required if name not in idx]
    if missing:
        raise ValueError(
            f"ETF_info.xlsx is missing expected column(s): {', '.join(missing)}"
        )

    metadata: dict[str, dict[str, str]] = {}
    for row in rows:
        if len(row) <= max(idx.values()):
            continue
        ticker = row[idx["ticker"]]
        category = row[idx["category"]]
        fund_family = row[idx["fundFamily"]]
        if not ticker or not category or not fund_family:
            continue  # incomplete record - skip rather than guess
        metadata[str(ticker)] = {
            "category": str(category),
            "fund_family": str(fund_family),
            "issuer_tier": classify_issuer_tier(str(fund_family)),
        }
    return metadata


class FundamentalMetadataDataService:
    """Serves PRICE_VOLUME (delegated) and ETF_METADATA (local fixture)."""

    def __init__(self, etf_info_path: Path | str = DEFAULT_ETF_INFO_PATH) -> None:
        self._etf_info_path = Path(etf_info_path)
        self._price_service = YFinanceDataService()
        self._metadata_cache: dict[str, dict[str, str]] | None = None

    async def fetch(self, request: DataRequest) -> DataResponse:
        artifacts: list[DataArtifact] = []
        unavailable: list[str] = []

        wants_price = (
            not request.categories or DataCategory.PRICE_VOLUME in request.categories
        )
        wants_metadata = (
            not request.categories or DataCategory.ETF_METADATA in request.categories
        )

        if wants_price:
            price_response = await self._price_service.fetch(request)
            artifacts.extend(price_response.artifacts)
            unavailable.extend(price_response.unavailable_fields)

        if wants_metadata:
            metadata_artifact = self._build_metadata_artifact(request)
            if metadata_artifact is not None:
                artifacts.append(metadata_artifact)
            else:
                unavailable.append("category")
                unavailable.append("fund_family")

        return DataResponse(
            response_id=f"{request.request_id}.response",
            request_id=request.request_id,
            lineage=request.lineage,
            as_of_date=request.as_of_date,
            complete=bool(artifacts) and not unavailable,
            artifacts=artifacts,
            unavailable_fields=unavailable,
            limitations=list(_METADATA_LIMITATIONS),
        )

    def _build_metadata_artifact(self, request: DataRequest) -> DataArtifact | None:
        if self._metadata_cache is None:
            if not self._etf_info_path.exists():
                return None
            self._metadata_cache = _load_etf_metadata(self._etf_info_path)

        symbols = (
            request.asset_universe
            if isinstance(request.asset_universe, list) and request.asset_universe
            else list(DEFAULT_UNIVERSE)
        )
        payload = {
            symbol: self._metadata_cache[symbol]
            for symbol in symbols
            if symbol in self._metadata_cache
        }
        if not payload:
            return None

        now = datetime.now(timezone.utc)
        return DataArtifact(
            artifact_id=f"{request.request_id}.etf_metadata",
            category=DataCategory.ETF_METADATA,
            description="ETF category and fund family from ETF_info.xlsx.",
            data_reference=f"static_xlsx::{self._etf_info_path}::etf_metadata",
            schema_fields=["ticker", "category", "fund_family", "issuer_tier"],
            asset_scope=sorted(payload),
            provenance=[
                DataProvenance(
                    provenance_id=f"{request.request_id}.etf_metadata.provenance",
                    provider="etf_info_xlsx",
                    source_reference=str(self._etf_info_path),
                    retrieved_at=now,
                    point_in_time_verified=False,
                    notes=list(_METADATA_LIMITATIONS),
                )
            ],
            analysis_payload=payload,
            limitations=list(_METADATA_LIMITATIONS),
        )


__all__ = ["DEFAULT_ETF_INFO_PATH", "FundamentalMetadataDataService"]
